# -*- coding: utf-8 -*-
"""Parser lista de precios -> COLORS_DB.
REGLA DE ORO (bug julio 2026): una fila CON precio es SIEMPRE un color,
nunca un encabezado. El chequeo de seccion solo corre sobre filas sin precio.
"""
import openpyxl, re, sys, json

BRAND_MARKERS = [
    ('GUIDONI QUARTZ',                'Guidoni',    'USD'),
    ('LINEA  STELLAR',                'Guidoni',    'USD'),
    ('LINEA STELLAR',                 'Guidoni',    'USD'),
    ('TRAVERTINOS (USD)',             'Marmol',     'USD'),
    ('MARMOLES IMPORTADOS',           'Marmol',     'USD'),
    ('MARMOLES & TRAVERTINOS',        'Marmol',     'USD'),
    ('MARMETAS GRANITOS',             'Marmol',     'USD'),
    ('GRANITOS NACIONALES',           'Granito_n',  'ARS'),
    ('GRANITOS IMPORTADOS',           'Granito_i',  'USD'),
    ('GRANITOS VERDES',               'Granito_i',  'USD'),
    ('GRANITOS BLANCO',               'Granito_i',  'USD'),
    ('GRANITOS GRIS',                 'Granito_i',  'USD'),
    ('MARMOLES NEGROS',               'Marmol',     'USD'),
    ('MARMOLES MARRONES',             'Marmol',     'USD'),
    ('XTONE',                         'Xtone',      'USD'),
    ('CUARCITAS IMPORTADOS',          'Cuarcita',   'USD'),
    ('PURASTONE',                     'Pura',       'USD'),
    ('DEKTON',                        'Dekton',     'USD'),
    ('PRIMA  DE STEFANO',             'Prima',      'USD'),
    ('PRIMA DE STEFANO',              'Prima',      'USD'),
    ('SILESTONE',                     'Silestone',  'USD'),
    ('NEOLITH',                       'Neolith',    'USD'),
    ('SUPRASTONE',                    'Suprastone', 'USD'),
    ('PIEDRAS NATURALES EXCLUSIVA',   'Marmol',     'USD'),
    ('MARMOLES Y CALIZAS',            'Marmol',     'USD'),
    ('MARMOL BLANCOS VETEADOS',       'Marmol',     'USD'),
    ('STEFANO -TERRAZO',              'Marmol',     'USD'),
    ('STEFANO-TERRAZO',               'Marmol',     'USD'),
    ('CUARCITAS  Y GRANITOS EXOTICOS','Cuarcita',   'USD'),
    ('CUARCITAS Y GRANITOS EXOTICOS', 'Cuarcita',   'USD'),
    ('CUARCITAS CANTERA DEL MUNDO',   'Cuarcita',   'USD'),
    ('SERVICIOS',                     None,         None),
]

def norm(s):
    return ' '.join(str(s).split())

def clean_name(n, brand):
    n = norm(n)
    # el xlsx a veces trae el nombre de la marca adelante
    if brand == 'Silestone':
        n = re.sub(r'^silestone\s+', '', n, flags=re.I)
    return n.strip()

def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    brand, cur = None, 'USD'
    out, skipped = [], []
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
        a = r[0]; b = r[1] if len(r) > 1 else None
        name = norm(a) if a is not None else ''
        price = b if isinstance(b, (int, float)) else None

        if price is None:
            # solo aca puede haber encabezado
            up = name.upper()
            for marker, br, cu in BRAND_MARKERS:
                if marker in up:
                    brand = br
                    if cu: cur = cu
                    break
            continue

        if not name or price <= 0:
            skipped.append((i, name, price)); continue
        if brand is None:
            skipped.append((i, name, price)); continue
        out.append({'row': i, 'n': clean_name(name, brand), 'm': brand,
                    'p': round(float(price), 2), 'c': cur})
    return out, skipped

if __name__ == '__main__':
    items, skipped = parse(sys.argv[1])
    from collections import Counter
    print('items:', len(items))
    print(Counter(x['m'] for x in items).most_common())
    print('skipped:', len(skipped))
    for s in skipped[:20]: print('  skip', s)
    json.dump(items, open('parsed.json','w'), ensure_ascii=False)
